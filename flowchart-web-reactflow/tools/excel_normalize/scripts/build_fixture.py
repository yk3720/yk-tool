"""入力用 Excel fixture（テンプレ v0.1）を生成する。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from excel_normalize.constants import FLOW_HEADERS, KOSEI_HEADERS, KOSEI_SHEET

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "fixtures"

INTERNAL_CODE = "Z00001"
DISPLAY_NAME = "プレス機B"

KOSEI_ROWS = [
    (INTERNAL_CODE, DISPLAY_NAME, "供給ユニット", "取出"),
    (INTERNAL_CODE, DISPLAY_NAME, "供給ユニット", "供給"),
    (INTERNAL_CODE, DISPLAY_NAME, "収納ユニット", "取出"),
    (INTERNAL_CODE, DISPLAY_NAME, "収納ユニット", "収納"),
]

FLOW_SAMPLES: dict[str, list[list[str]]] = {
    "取出": [
        ["10", "端子", "20", "", "1", "0", "開始", "", "", ""],
        ["20", "処理", "30", "", "2", "0", "ワーク取出", "", "", ""],
        ["30", "端子", "", "", "3", "0", "終了", "", "", ""],
    ],
    "供給": [
        ["10", "端子", "20", "", "1", "0", "開始", "", "", ""],
        ["20", "処理", "30", "", "2", "0", "供給実行", "", "", ""],
        ["30", "端子", "", "", "3", "0", "終了", "", "", ""],
    ],
    "収納": [
        ["10", "端子", "20", "", "1", "0", "開始", "", "", ""],
        ["20", "処理", "30", "", "2", "0", "収納実行", "", "", ""],
        ["30", "端子", "", "", "3", "0", "終了", "", "", ""],
    ],
}


def _add_flow_table(
    ws,
    *,
    table_name: str,
    start_col: int,
    start_row: int,
    data_rows: list[list[str]],
) -> None:
    headers = list(FLOW_HEADERS)
    row_count = 1 + len(data_rows)
    end_col = start_col + len(headers) - 1
    end_row = start_row + row_count - 1

    for c, header in enumerate(headers):
        ws.cell(start_row, start_col + c, header)
    for r_idx, data in enumerate(data_rows, start=1):
        for c_idx, value in enumerate(data):
            ws.cell(start_row + r_idx, start_col + c_idx, value)

    start_cell = ws.cell(start_row, start_col).coordinate
    end_cell = ws.cell(end_row, end_col).coordinate
    ref = f"{start_cell}:{end_cell}"

    safe_name = table_name.replace(" ", "_")
    table = Table(displayName=safe_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws_kosei = wb.active
    ws_kosei.title = KOSEI_SHEET
    ws_kosei.append(list(KOSEI_HEADERS))
    for row in KOSEI_ROWS:
        ws_kosei.append(list(row))

    ws_supply = wb.create_sheet("供給ユニット")
    _add_flow_table(
        ws_supply,
        table_name="供給_取出",
        start_col=1,
        start_row=1,
        data_rows=FLOW_SAMPLES["取出"],
    )
    _add_flow_table(
        ws_supply,
        table_name="供給_供給",
        start_col=13,
        start_row=1,
        data_rows=FLOW_SAMPLES["供給"],
    )

    ws_storage = wb.create_sheet("収納ユニット")
    _add_flow_table(
        ws_storage,
        table_name="収納_取出",
        start_col=1,
        start_row=1,
        data_rows=FLOW_SAMPLES["取出"],
    )
    _add_flow_table(
        ws_storage,
        table_name="収納_収納",
        start_col=13,
        start_row=1,
        data_rows=FLOW_SAMPLES["収納"],
    )

    return wb


def main() -> None:
    FIXTURES.mkdir(parents=True, exist_ok=True)
    out = FIXTURES / "input-device-z00001.xlsx"
    build_workbook().save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
