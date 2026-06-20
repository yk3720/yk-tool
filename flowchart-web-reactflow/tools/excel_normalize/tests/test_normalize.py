from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from excel_normalize.constants import KOSEI_HEADERS, KOSEI_SHEET
from excel_normalize.normalize import normalize_workbook

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = ROOT / "fixtures"
TEMPLATES = ROOT / "templates"
INPUT_XLSX = FIXTURES / "input-device-z00001.xlsx"
TEMPLATE_XLSX = TEMPLATES / "入力用テンプレ_v0.2.xlsx"
A0001_MASTER_XLSX = FIXTURES / "devices" / "A0001_塗布装置" / "マスター.xlsx"


@pytest.fixture(scope="session")
def input_xlsx() -> Path:
    if not INPUT_XLSX.is_file():
        pytest.skip("fixture 未生成 — scripts/build_fixture.py を実行してください")
    return INPUT_XLSX


@pytest.fixture(scope="session")
def template_xlsx() -> Path:
    if not TEMPLATE_XLSX.is_file():
        pytest.skip("テンプレ未生成 — scripts/build_template.py を実行してください")
    return TEMPLATE_XLSX


@pytest.fixture(scope="session")
def a0001_master_xlsx() -> Path:
    if not A0001_MASTER_XLSX.is_file():
        pytest.skip("A0001 マスター未生成 — npm run excel:a0001:build")
    return A0001_MASTER_XLSX


def test_normalize_produces_bundle(input_xlsx: Path) -> None:
    result = normalize_workbook(input_xlsx)
    bundle = result.bundle

    assert bundle["internal_code"] == "Z00001"
    assert bundle["display_name"] == "プレス機B"
    assert len(bundle["units"]) == 2
    assert len(bundle["flows"]) == 4

    for flow in bundle["flows"]:
        assert flow["payload"]["schema"] == "table-10col-v1"
        assert len(flow["payload"]["table"]) >= 1
        assert len(flow["payload"]["table"][0]) == 10


def test_template_normalizes(template_xlsx: Path) -> None:
    result = normalize_workbook(template_xlsx)
    bundle = result.bundle

    assert bundle["internal_code"] == "NEW-001"
    assert bundle["display_name"] == "（装置名）"
    assert len(bundle["units"]) == 2
    assert len(bundle["flows"]) == 4
    unit_labels = {u["label"] for u in bundle["units"]}
    assert unit_labels == {"供給ユニット", "加工ユニット"}


def test_a0001_master_normalizes(a0001_master_xlsx: Path) -> None:
    result = normalize_workbook(a0001_master_xlsx)
    bundle = result.bundle

    assert bundle["internal_code"] == "A0001"
    assert bundle["display_name"] == "塗布装置"
    assert len(bundle["units"]) == 2
    assert len(bundle["flows"]) == 4

    toridashi = next(f for f in bundle["flows"] if f["module_label"] == "取出")
    assert toridashi["unit_label"] == "供給ユニット"
    assert toridashi["payload"]["table"][1][6] == "ワーク取出"

    module_labels = [m["label"] for u in bundle["units"] for m in u["modules"]]
    assert module_labels == ["取出", "供給", "プレス", "離脱"]


def test_normalize_writes_json_snapshot(input_xlsx: Path, tmp_path: Path) -> None:
    result = normalize_workbook(input_xlsx)
    out = tmp_path / "import.json"
    out.write_text(
        json.dumps(result.bundle, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    assert out.stat().st_size > 0


def test_missing_kosei_sheet(tmp_path: Path) -> None:
    wb = Workbook()
    wb.active.title = "供給ユニット"
    path = tmp_path / "no-kosei.xlsx"
    wb.save(path)
    with pytest.raises(ValueError, match="構成"):
        normalize_workbook(path)


def test_table_name_mismatch(tmp_path: Path) -> None:
    if not INPUT_XLSX.is_file():
        pytest.skip("fixture 未生成")

    wb = load_workbook(INPUT_XLSX)
    ws = wb["供給ユニット"]
    for table in list(ws.tables.values()):
        if table.name == "供給_取出":
            table.name = "unknown_action"
            break
    path = tmp_path / "bad-table.xlsx"
    wb.save(path)

    with pytest.raises(ValueError, match="一致しません"):
        normalize_workbook(path)


def _minimal_kosei_wb(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = KOSEI_SHEET
    ws.append(list(KOSEI_HEADERS))
    ws.append(["Z00001", "プレス機B", "供給ユニット", "取出"])
    path = tmp_path / "missing-table.xlsx"
    wb.save(path)
    return path


def test_missing_flow_table(tmp_path: Path) -> None:
    path = _minimal_kosei_wb(tmp_path)
    with pytest.raises(ValueError, match="供給ユニット"):
        normalize_workbook(path)
